import pandas as pd
import os
from openpyxl import load_workbook
from test1 import postProcess_CBSE_SSC,postProcess_State_Inter,postProcess_State_SSC


Prompt_SSC_State=""" extract the data and give 
Board
Candidate Name
Father Name
Mother Name
Roll Number
Date of Birth
School Name
Medium
all Subjects
CGPA
Date of Issue
Identification marks

note: for each  subject {Subject Name,Grade} take only 6 subjects and subject name must not be like {'First Language','Second Language','Third Language','Fourth Language','Fifth Language','Sixth Language'} must be like specific subject name

give just dictionary no other stuff
"""

Prompt_SSC_CBSE=""" extract the data and give 
Board
Candidate Name
Father Name
Mother Name
Roll Number
Date of Birth
School Name
all Subjects
Result
Date of Issue

note: for each  Subject {Subject Code,Subject Name,Grade} add only 5 subjects in subjects array
take a separate dict for adiitional subject that is 6th subject if there is no fill this with null values but same keys
give just dictionary no other stuff
"""

Prompt_SSC_ICSE=""" extract the data and give 
Board
Candidate Name
Father Name
Mother Name
Roll Number
Date of Birth
School Name
Medium
Exam Date
all subjects
CGPA
Date of Issue
Identification marks

note: for each  subject {Name,Grade}

give just dictionary no other stuff
"""

Prompt_Inter_State=""" extract the data and give 
Board
Candidate Name
Father Name
Mother Name
Roll Number
Total Marks
all Subjects
Practiacls
Date of Issue

note: for each  Subject {Subject Name, 1st Yr Marks, 2nd Yr Marks} take only 6 subjects start with part 1 subjects
for each  Practical {Subject Name, Marks}


give just dictionary no other stuff
"""

Prompt_Inter_CBSE=""" extract the data and give 
Board
Candidate Name
Father Name
Mother Name
Roll Number
Date of Birth
School Name
Medium
Exam Date
all subjects
CGPA
Date of Issue
Identification marks

note: for each  subject {Name,Grade}

give just dictionary no other stuff
"""

Prompt_Inter_ICSE=""" extract the data and give 
Board
Candidate Name
Father Name
Mother Name
Roll Number
Date of Birth
School Name
Medium
Exam Date
all subjects
CGPA
Date of Issue
Identification marks

note: for each  subject {Name,Grade}

give just dictionary no other stuff
"""

General_Prompt=""" extract the data and give 
Board
Candidate Name 
Father Name
Mother Name
Roll Number
Date of Birth

give just dictionary no other stuff"""



def process_certificate(record,file_path):
    flat_data = {
        "Board": record["Board"],
        "Candidate Name": record["Candidate Name"],
        "Father Name": record.get("Father Name", None),
        "Mother Name": record["Mother Name"],
        "Roll Number": int(record["Roll Number"]),
        "Date of Birth": record["Date of Birth"],
    }
    flat_data["address"]=file_path.resolve()
    flat_data["File path"]=record["Candidate Name"]
    
    return flat_data


def append_to_excel(new_record, excel_path, certificate_path, Standard, Board):
    new_data = pd.DataFrame([new_record])
    if os.path.exists(excel_path):
        existing_data = pd.read_excel(excel_path)
        updated_data = pd.concat([existing_data, new_data], ignore_index=True)
        updated_data = updated_data.drop_duplicates(subset=['Roll Number'], keep='first')  # Keeps the latest one (or 'first' to keep the first)

    else:
        updated_data = new_data

    updated_data.to_excel(excel_path, index=False)
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # Add hyperlinks to the 'File path' column (assuming it's the last column)
    file_path_column_index = None
    for idx, col_name in enumerate(updated_data.columns, 1):  # 1-based index
        if col_name == "File path":
            file_path_column_index = idx
            break
    
    if file_path_column_index is not None:
        # Add hyperlinks to the 'File path' column (assuming it's found)
        for row in range(2, len(updated_data) + 2):  # Start from row 2 to avoid header
            file_path = ws.cell(row=row, column=file_path_column_index-1).value
            # file_path=path_certificate
            if file_path:
                # Add the hyperlink to the cell
                ws.cell(row=row, column=file_path_column_index).hyperlink = file_path
                ws.cell(row=row, column=file_path_column_index).style = 'Hyperlink'
    
    # Save the workbook with the hyperlinks
    wb.save(excel_path)
    return excel_path
